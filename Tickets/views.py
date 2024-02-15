from django.shortcuts import render
from rest_framework.decorators import api_view
# Create your views here.
import openpyxl
from .models import State,UserMaster,TaskMaster
from django.http import HttpResponse
from django.db import transaction
from .serializers import TaskMasterSerializer
from django.http import JsonResponse
import io

from django.core.paginator import Paginator,PageNotAnInteger,EmptyPage


# def parser(request):
#     if request.method=="GET":
#         file="C:\\Users\Maqsood\\Desktop\\Django_Todo\\Project_Minute\\states.xlsx"
#         workbook = openpyxl.load_workbook(file)
#         sheet = workbook.active
#         for row in sheet.iter_rows(min_row=2,values_only=True):
#             print(row)
#             states,state_code=row
#             d_state=State.objects.filter(state_name=states).first()
#             if d_state:
#                 continue
#             update=State(state_name=states,state_code=state_code)
#             update.save()
#         return HttpResponse("updated")
    
# def userMasterView(request):
#     if request.method=="GET":
#         file="C:\\Users\Maqsood\\Desktop\\Django_Todo\\Project_Minute\\MOCK_DATA.xlsx"
#         workbook = openpyxl.load_workbook(file)
#         sheet = workbook.active
#         a=0
#         state = State.objects.all()
#         state_length = len(state)
#         for row in sheet.iter_rows(min_row=2,values_only=True):
#             name,email,password,phone,city,gender,joining_date=row
#             check=UserMaster.objects.filter(name=name,email=email,phone=phone,joining_date=joining_date)
#             if check:
#                 continue
#             print(state[0].id)
#             data_updating=UserMaster(name=name,email=email,password=password,phone=phone,
#             city=city, gender=gender, joining_date=joining_date, state=state[a])
#             data_updating.save()
#             a+=1
#             if a==(state_length):
#                 a=0
#             print(a)
#         return HttpResponse("Done")
    
# def UserMasterBulkDataView(request):
#     if request.method =="GET":
#         file="C:\\Users\Maqsood\\Desktop\\Django_Todo\\Project_Minute\\MOCK_DATA2.xlsx"
#         workbook = openpyxl.load_workbook(file)
#         sheet = workbook.active
#         a=0
#         state = State.objects.all()
#         state_length = len(state)
#         user_data=[]
#         for row in sheet.iter_rows(min_row=2,values_only=True):
#             name,email,password,phone,city,gender,joining_date=row
#             check=UserMaster.objects.filter(name=name,email=email,phone=phone,joining_date=joining_date)
#             if check:
#                 continue
#             data_updating=UserMaster(name=name,email=email,password=password,phone=phone,
#             city=city, gender=gender, joining_date=joining_date, state=state[a])
#             user_data += [data_updating]
#             a+=1
#             if a==(state_length):
#                 a=0
#         with transaction.atomic():
#             UserMaster.objects.bulk_create(user_data)
#         return HttpResponse("Done")

# assign task to Users
def AssignTasks(request):
    add_state=State.objects.all().order_by('state_name').distinct('state_name')
    if request.method=="POST":
        alert=True
        task_id=request.POST.get('taskid')
        title=request.POST.get('title')
        description=request.POST.get('description')
        assign_date=request.POST.get('assign_date')
        state=request.POST.get('state')
        assign_to=request.POST.get('assign_to')
        
        
        check_state=State.objects.filter(state_name=state)
        try:
            assign_to=int(assign_to)
        except Exception as e:
            print(e)
    
        if type(assign_to)==int and check_state:
            check_assign_to=UserMaster.objects.filter(id=assign_to)
            check_duplicate_task=TaskMaster.objects.filter(task_id=task_id,title=title,description=description,user_id=check_assign_to[0]).first()
            if check_duplicate_task:
                return render(request,'AssignTasks.html',{'date':check_duplicate_task.assign_date,'user':check_duplicate_task.user_id,'states':add_state})
            
            data=TaskMaster(task_id=task_id,title=title,
                            description=description,assign_date=assign_date,
                            state=check_state[0],user_id=check_assign_to[0])
            data.save()
            return render(request,"AssignTasks.html",{'states':add_state,'alert':alert})
        
        elif type(assign_to)==str and check_state:
            check_name_assign_to=UserMaster.objects.filter(name=assign_to)
            check_duplicate_task=TaskMaster.objects.filter(task_id=task_id,title=title,description=description,user_id=check_name_assign_to[0]).first()
            if check_duplicate_task:
                return render(request,'AssignTasks.html',{'date':check_duplicate_task.assign_date,'user':check_duplicate_task.user_id,'states':add_state})
            
            data=TaskMaster(task_id=task_id,title=title,
                            description=description,assign_date=assign_date,
                            state=check_state[0],user_id=check_name_assign_to[0])
            data.save()
            return render(request,"AssignTasks.html",{'states':add_state,'alert':alert,'states':add_state})
        else:
            alert=False
            return render(request,"AssignTasks.html",{'alert':alert,'states':add_state})
    return render(request,"AssignTasks.html",{'states':add_state})

def ViewTasks(request,id=None):
    items_per_page = 5
    page_number = request.GET.get('page', 1)
    search_query = request.GET.get('search')
    if search_query:
        tasks = TaskMaster.objects.filter(title__icontains=search_query).order_by('id')
    else:
        tasks = TaskMaster.objects.all().order_by('id')
    paginator = Paginator(tasks, items_per_page)
    paginated_data = paginator.get_page(page_number)
    total_pages = paginator.num_pages
    data={'data': paginated_data,'lastpage':total_pages,'pageNumbers':[n+1 for n in range(total_pages)],
                                                     'search_query':search_query}
    
    return render(request, 'ViewAssignedTask.html',data)

from django.shortcuts import get_object_or_404, redirect

def DeleteTask(request, id):
    task = get_object_or_404(TaskMaster, id=id)
    task.delete()
    return redirect('view-tasks')


def update(request,id):
    
    task = get_object_or_404(TaskMaster,id=id)
    if request.method=="POST":
        print(task.assign_date)
        if id:
            task_id=request.POST.get('taskid')
            title=request.POST.get('title')
            description = request.POST.get('description')
            assign_date = request.POST.get('assign_date')
            if assign_date:
                task.assign_date=assign_date
            task.task_id=task_id
            task.title=title
            task.description=description
            

            task.save()
            return redirect('view-tasks')
    return render(request, 'update.html', {'task': task})


def homepage(request):
    return render(request,'index.html')

def aboutpage(request):
    return render(request,'about.html')

def adduser(request):
    states=State.objects.all().order_by('state_name').distinct('state_name')
    alert=True
    if request.method=="POST":
        name=request.POST.get('name')
        email=request.POST.get('email')
        password=request.POST.get('password')
        phone=request.POST.get('phone')
        city=request.POST.get('city')
        gender=request.POST.get('gender')
        joining_date=request.POST.get('joining_date')
        state=request.POST.get('state')
        check_state=State.objects.filter(state_name=state)
        
        check_duplicate=UserMaster.objects.filter(name=name,email=email,phone=phone,city=city)
        if check_duplicate:
            added=False
            return render(request,'adduser.html',{'states':states,'added':added})
        if check_state:
            data=UserMaster(name=name,email=email,password=password,phone=phone,city=city,
                            gender=gender,joining_date=joining_date,state=check_state[0])
            data.save()
            added=True
        return render(request,'adduser.html',{'states':states,'added':added})
    return render(request,'adduser.html',{'states':states,'alert':alert})

def addmultipleusers(request):
    alert=True
    if request.method =="POST":
        file=request.FILES['excelFile']
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        user_data=[]
        added=False
        duplicate_users=0
        not_added=0
        for row in sheet.iter_rows(min_row=2,values_only=True):
            if len(row)<8 or len(row)>8:
                return render(request,"multiple_user_error.html")
            name,email,password,phone,city,gender,joining_date,state=row
            check=UserMaster.objects.filter(name=name,email=email,password=password,phone=phone,city=city)
            if check:
                duplicate_users+=1
                continue
            check_states=State.objects.filter(state_name=state)
            if check_states:
                data_updating=UserMaster(name=name,email=email,password=password,phone=phone,
                city=city, gender=gender, joining_date=joining_date, state=check_states[0])
                user_data += [data_updating]
            else:
                not_added+=1
        with transaction.atomic():
            UserMaster.objects.bulk_create(user_data)
            added = True
            if added and duplicate_users:
                return render(request,'addmultipleusers.html',{'added':added,'duplicate_user':duplicate_users,'not_added':not_added})
            return render(request,'addmultipleusers.html',{'added':added})
    return render(request,'addmultipleusers.html',{'alert':alert})


def addState(request):
    if request.method == "POST":
        state_name = request.POST.get('state_name')
        state_code = request.POST.get('state_code')
        check_state = State.objects.filter(state_name=state_name, state_code=state_code)
        if check_state:
            alert = False
            return render(request, 'addstates.html', {'alert': alert})
        obj = State(state_name=state_name, state_code=state_code)
        obj.save()
        alert=True
        return render(request, 'addstates.html', {'alert': alert})
    return render(request, 'addstates.html')

def addMultipleState(request):
    alert=True
    if request.method=="POST":
        file=request.FILES['excelFile']
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        added=False
        duplicate_users=0
        state_data=[]
        for row in sheet.iter_rows(min_row=2,values_only=True):
            if len(row)>2 or len(row)<2:
                return render(request,'multiple_states_error.html')
            states,state_code=row
            d_state=State.objects.filter(state_name=states,state_code=state_code).first()
            if d_state:
                duplicate_users+=1
                continue
            add_states=State(state_name=states,state_code=state_code)
            state_data+=[add_states]
        with transaction.atomic():
            State.objects.bulk_create(state_data)
            added = True
            if added and duplicate_users:
                return render(request,'addmultiplestates.html',{'added':added,'duplicate_user':duplicate_users})
            return render(request,'addmultiplestates.html',{'added':added})
    return render(request,'addmultiplestates.html',{'alert':alert})




#page not found
# def handler404(request,unwanted=None,second=None):
#     return render(request, '404.html', status=404)
